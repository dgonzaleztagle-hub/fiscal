from datetime import datetime, timezone
import hashlib
from io import BytesIO

from openpyxl import load_workbook

from completo_dte.application import IssueInvoiceService, MonthlyReportBuilder
from test_issue_invoice import setup_invoice
from test_received_ledger import received_document


def test_monthly_report_combines_sales_purchases_and_exports_hashed_csv(tmp_path) -> None:
    ledger, caf, _caf_id, credential, command = setup_invoice(tmp_path)
    sale = IssueInvoiceService(
        ledger=ledger,
        resolve_caf=lambda _requested: caf,
        resolve_credential=lambda _tenant, _rut: credential,
        validate_signed_dte=lambda _document: None,
        clock=lambda: datetime(2026, 7, 10, 15, 30, tzinfo=timezone.utc),
    ).issue(command)
    purchase = ledger.import_received_document(
        tenant_id="tenant-a", document=received_document(), source="upload"
    )
    builder = MonthlyReportBuilder()
    report = builder.build(
        year=2026, month=7, outgoing=[sale], received=[purchase]
    )
    artifact = builder.csv(report)

    assert len(report.rows) == 2
    assert report.sales_total == 11900
    assert report.purchases_total == 21420
    assert artifact.filename == "completo-fiscal-2026-07.csv"
    assert artifact.content.startswith(b"\xef\xbb\xbf")
    assert artifact.sha256 == hashlib.sha256(artifact.content).hexdigest()
    assert b"xml_sha256" in artifact.content
    xlsx = builder.xlsx(report)
    workbook = load_workbook(BytesIO(xlsx.content), data_only=True)
    assert workbook.sheetnames == ["Resumen", "Documentos"]
    assert workbook["Resumen"]["B2"].value == 11900
    assert workbook["Resumen"]["B3"].value == 21420
    assert workbook["Documentos"].max_row == 3
    assert xlsx.sha256 == hashlib.sha256(xlsx.content).hexdigest()
    pdf = builder.pdf(report)
    assert pdf.content.startswith(b"%PDF-")
    assert pdf.sha256 == hashlib.sha256(pdf.content).hexdigest()


def test_monthly_report_excludes_other_periods(tmp_path) -> None:
    ledger, caf, _caf_id, credential, command = setup_invoice(tmp_path)
    sale = IssueInvoiceService(
        ledger=ledger,
        resolve_caf=lambda _requested: caf,
        resolve_credential=lambda _tenant, _rut: credential,
        validate_signed_dte=lambda _document: None,
        clock=lambda: datetime(2026, 7, 10, 15, 30, tzinfo=timezone.utc),
    ).issue(command)
    report = MonthlyReportBuilder().build(
        year=2026, month=8, outgoing=[sale], received=[]
    )
    assert report.rows == ()
