"""Reporte mensual reproducible derivado de XML inmutables."""

from dataclasses import dataclass
import csv
import hashlib
from io import BytesIO, StringIO

from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from completo_dte.infrastructure import (
    FiscalDocumentRecord,
    ReceivedFiscalDocumentRecord,
)


@dataclass(frozen=True)
class MonthlyDocumentRow:
    direction: str
    document_type: int
    folio: int
    taxpayer_rut: str
    issued_on: str
    net: int
    exempt: int
    vat: int
    total: int
    xml_sha256: str


@dataclass(frozen=True)
class MonthlyFiscalReport:
    period: str
    rows: tuple[MonthlyDocumentRow, ...]

    @property
    def sales_total(self) -> int:
        return sum(row.total for row in self.rows if row.direction == "sale")

    @property
    def purchases_total(self) -> int:
        return sum(row.total for row in self.rows if row.direction == "purchase")


@dataclass(frozen=True)
class ExportArtifact:
    filename: str
    media_type: str
    content: bytes
    sha256: str


class MonthlyReportBuilder:
    def build(
        self,
        *,
        year: int,
        month: int,
        outgoing: list[FiscalDocumentRecord],
        received: list[ReceivedFiscalDocumentRecord],
    ) -> MonthlyFiscalReport:
        period = f"{year:04d}-{month:02d}"
        rows = []
        for record in outgoing:
            values = _xml_values(record.signed_xml)
            if values["issued_on"].startswith(period):
                rows.append(
                    MonthlyDocumentRow(
                        "sale",
                        record.document_type,
                        record.folio,
                        record.taxpayer_rut,
                        values["issued_on"],
                        values["net"],
                        values["exempt"],
                        values["vat"],
                        values["total"],
                        record.xml_sha256,
                    )
                )
        for record in received:
            if record.issued_on.startswith(period):
                values = _xml_values(record.signed_xml)
                rows.append(
                    MonthlyDocumentRow(
                        "purchase",
                        record.document_type,
                        record.folio,
                        record.issuer_rut,
                        record.issued_on,
                        values["net"],
                        values["exempt"],
                        values["vat"],
                        record.total,
                        record.xml_sha256,
                    )
                )
        return MonthlyFiscalReport(
            period,
            tuple(
                sorted(
                    rows, key=lambda row: (row.direction, row.document_type, row.folio)
                )
            ),
        )

    def csv(self, report: MonthlyFiscalReport) -> ExportArtifact:
        output = StringIO(newline="")
        writer = csv.writer(output, lineterminator="\n")
        writer.writerow(
            [
                "direction",
                "document_type",
                "folio",
                "taxpayer_rut",
                "issued_on",
                "net",
                "exempt",
                "vat",
                "total",
                "xml_sha256",
            ]
        )
        for row in report.rows:
            writer.writerow(
                [
                    row.direction,
                    row.document_type,
                    row.folio,
                    row.taxpayer_rut,
                    row.issued_on,
                    row.net,
                    row.exempt,
                    row.vat,
                    row.total,
                    row.xml_sha256,
                ]
            )
        content = output.getvalue().encode("utf-8-sig")
        return ExportArtifact(
            f"completo-fiscal-{report.period}.csv",
            "text/csv; charset=utf-8",
            content,
            hashlib.sha256(content).hexdigest(),
        )

    def xlsx(self, report: MonthlyFiscalReport) -> ExportArtifact:
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Resumen"
        summary.append(["Completo Fiscal", report.period])
        summary.append(["Ventas", report.sales_total])
        summary.append(["Compras", report.purchases_total])
        summary["A1"].font = Font(bold=True, size=14)
        summary.column_dimensions["A"].width = 22
        summary.column_dimensions["B"].width = 18
        for cell in (summary["B2"], summary["B3"]):
            cell.number_format = "$#,##0"

        sheet = workbook.create_sheet("Documentos")
        headers = [
            "Dirección",
            "Tipo",
            "Folio",
            "RUT",
            "Fecha",
            "Neto",
            "Exento",
            "IVA",
            "Total",
            "SHA-256 XML",
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="17233B")
        for row in report.rows:
            sheet.append(
                [
                    row.direction,
                    row.document_type,
                    row.folio,
                    row.taxpayer_rut,
                    row.issued_on,
                    row.net,
                    row.exempt,
                    row.vat,
                    row.total,
                    row.xml_sha256,
                ]
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in ("F", "G", "H", "I"):
            for cell in sheet[column][1:]:
                cell.number_format = "$#,##0"
        widths = [13, 10, 12, 16, 13, 14, 14, 14, 14, 66]
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[chr(64 + index)].width = width
        output = BytesIO()
        workbook.save(output)
        content = output.getvalue()
        return ExportArtifact(
            f"completo-fiscal-{report.period}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content,
            hashlib.sha256(content).hexdigest(),
        )

    def pdf(self, report: MonthlyFiscalReport) -> ExportArtifact:
        output = BytesIO()
        document = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=12 * mm,
            leftMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
        )
        styles = getSampleStyleSheet()
        story = [
            Paragraph(f"Completo Fiscal · Reporte {report.period}", styles["Title"]),
            Paragraph(
                f"Ventas: ${report.sales_total:,} · Compras: ${report.purchases_total:,}",
                styles["Normal"],
            ),
            Spacer(1, 6 * mm),
        ]
        data = [
            ["Dir.", "Tipo", "Folio", "RUT", "Fecha", "Neto", "Exento", "IVA", "Total"]
        ]
        for row in report.rows:
            data.append(
                [
                    row.direction,
                    row.document_type,
                    row.folio,
                    row.taxpayer_rut,
                    row.issued_on,
                    row.net,
                    row.exempt,
                    row.vat,
                    row.total,
                ]
            )
        table = Table(
            data,
            repeatRows=1,
            colWidths=[
                18 * mm,
                14 * mm,
                20 * mm,
                28 * mm,
                24 * mm,
                24 * mm,
                24 * mm,
                24 * mm,
                26 * mm,
            ],
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17233B")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9DEE7")),
                    ("ALIGN", (5, 1), (-1, -1), "RIGHT"),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#F7F8FA")],
                    ),
                ]
            )
        )
        story.append(table)
        document.build(story)
        content = output.getvalue()
        return ExportArtifact(
            f"completo-fiscal-{report.period}.pdf",
            "application/pdf",
            content,
            hashlib.sha256(content).hexdigest(),
        )


def _xml_values(payload: bytes) -> dict[str, str | int]:
    root = etree.fromstring(
        payload,
        etree.XMLParser(resolve_entities=False, no_network=True, load_dtd=False),
    )

    def value(name, default="0"):
        values = root.xpath("//*[local-name()=$name]/text()", name=name)
        return str(values[0]).strip() if len(values) == 1 else default

    return {
        "issued_on": value("FchEmis", ""),
        "net": int(value("MntNeto")),
        "exempt": int(value("MntExe")),
        "vat": int(value("IVA")),
        "total": int(value("MntTotal")),
    }
